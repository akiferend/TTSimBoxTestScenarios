import time
import os
import threading
from datetime import datetime
import customtkinter as ctk
from pymodbus.client import ModbusSerialClient

# Excel Raporlama Kütüphaneleri
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# PDF Raporlama Kütüphaneleri
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")


class HILTestApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Taytech TTSimBox - 8-Channel HIL Test Suite")
        self.geometry("850x880")

        self.com_port = "COM5"
        self.slave_id = 1

        # 8 POT KANALININ C++ TEST ENGINE İLE TAM UYUMLU SENARYOLARI
        self.test_scenarios = [
            {"id": 1, "name": "POT1 (AD5248 Ch0) - PT1000 0°C",   "type": "PT1000 Simülasyonu", "expected": "1000 Ω"},
            {"id": 2, "name": "POT2 (AD5248 Ch1) - PT1000 100°C", "type": "PT1000 Simülasyonu", "expected": "1385 Ω"},
            {"id": 3, "name": "POT3 (MCP4632 Ch3) - Direct Pot", "type": "Direnç Doğrulama",   "expected": "10 kΩ"},
            {"id": 4, "name": "POT4 (MCP4632 Ch4) - Direct Pot", "type": "Direnç Doğrulama",   "expected": "15 kΩ"},
            {"id": 5, "name": "POT5 (MCP4632 Ch5) - Direct Pot", "type": "Direnç Doğrulama",   "expected": "20 kΩ"},
            {"id": 6, "name": "POT6 (MCP4632 Ch6) - Direct Pot", "type": "Direnç Doğrulama",   "expected": "25 kΩ"},
            {"id": 7, "name": "POT7 (MCP4632 Ch7) - Direct Pot", "type": "Direnç Doğrulama",   "expected": "30 kΩ"},
            {"id": 8, "name": "POT8 (MCP4632 Ch8) - Direct Pot", "type": "Direnç Doğrulama",   "expected": "40 kΩ"},
        ]

        self.test_results = {}
        self.setup_ui()

    def setup_ui(self):
        self.header_label = ctk.CTkLabel(
            self, text="TTSimBox 8-CHANNEL HIL AUTOMATION TESTER",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        self.header_label.pack(pady=15)

        self.conn_frame = ctk.CTkFrame(self)
        self.conn_frame.pack(fill="x", padx=20, pady=5)

        self.port_label = ctk.CTkLabel(self.conn_frame, text="COM Port:", font=ctk.CTkFont(size=13))
        self.port_label.pack(side="left", padx=10, pady=8)

        self.port_entry = ctk.CTkEntry(self.conn_frame, width=80)
        self.port_entry.insert(0, self.com_port)
        self.port_entry.pack(side="left", padx=5, pady=8)

        self.run_all_btn = ctk.CTkButton(
            self.conn_frame, text="TÜM TEST SENARYOLARINI KOŞTUR",
            fg_color="#1D4ED8", hover_color="#1E40AF",
            command=lambda: self.start_thread(self.run_all_tests)
        )
        self.run_all_btn.pack(side="right", padx=10, pady=8)

        self.tests_frame = ctk.CTkFrame(self)
        self.tests_frame.pack(fill="x", padx=20, pady=10)

        self.single_buttons = []
        for idx, test in enumerate(self.test_scenarios):
            t_id = test["id"]
            btn = ctk.CTkButton(
                self.tests_frame,
                text=f"Test {t_id}: {test['name']} ({test['expected']})",
                fg_color="#374151",
                hover_color="#4B5563",
                command=lambda id_=t_id: self.start_thread(lambda: self.run_single_test(id_))
            )
            btn.grid(row=idx // 2, column=idx % 2, padx=10, pady=6, sticky="ew")
            self.single_buttons.append(btn)

        self.tests_frame.grid_columnconfigure(0, weight=1)
        self.tests_frame.grid_columnconfigure(1, weight=1)

        self.log_box = ctk.CTkTextbox(self, width=790, height=240, font=ctk.CTkFont(family="Consolas", size=12))
        self.log_box.pack(padx=20, pady=10)

        self.report_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.report_frame.pack(pady=10)

        self.excel_btn = ctk.CTkButton(
            self.report_frame, text="📊 EXCEL TEST RAPORU OLUŞTUR",
            fg_color="#107C41", hover_color="#0F6C38",
            state="disabled", command=self.generate_excel_report
        )
        self.excel_btn.pack(side="left", padx=10)

        self.pdf_btn = ctk.CTkButton(
            self.report_frame, text="📄 PDF TEST RAPORU OLUŞTUR",
            fg_color="#C0392B", hover_color="#A93226",
            state="disabled", command=self.generate_pdf_report
        )
        self.pdf_btn.pack(side="left", padx=10)

    def log(self, message):
        self.log_box.insert("end", message + "\n")
        self.log_box.see("end")

    def set_buttons_state(self, state):
        self.run_all_btn.configure(state=state)
        for btn in self.single_buttons:
            btn.configure(state=state)

    def enable_report_buttons(self):
        if self.test_results:
            self.excel_btn.configure(state="normal")
            self.pdf_btn.configure(state="normal")

    def start_thread(self, target_func):
        self.set_buttons_state("disabled")
        threading.Thread(target=target_func, daemon=True).start()

    def execute_modbus_test(self, client, test):
        t_id = test["id"]
        t_name = test["name"]

        self.log(f"--> Test {t_id} Koşturuluyor: {t_name} (Beklenen: {test['expected']})")

        try:
            # 1. Komutu ve Test ID'yi sıfırla/hazırla (slave yerine device_id kullanılıyor)
            client.write_register(0, 0, device_id=self.slave_id)
            time.sleep(0.05)

            client.write_register(1, t_id, device_id=self.slave_id)
            time.sleep(0.05)

            # 2. Testi Tetikle (REG_SYS_CMD = 1)
            client.write_register(0, 1, device_id=self.slave_id)
            
            # C++ tarafındaki 5 saniyelik zaman aşımını bekle
            time.sleep(5.5)

            # 3. Test Durumunu Oku (Reg 10-17 arası / Status Base)
            res = client.read_holding_registers(10, count=len(self.test_scenarios), device_id=self.slave_id)
            status_str = "READ_ERROR"

            if not res.isError():
                channel_statuses = res.registers  
                target_index = t_id - 1
                target_channel_status = channel_statuses[target_index]

                if target_channel_status == 2:  # STATUS_PASS (2)
                    status_str = "PASSED"
                    self.log(f"    [SONUÇ] Test {t_id} BAŞARILI (PASS)\n")
                else:
                    status_str = "FAILED"
                    self.log(f"    [SONUÇ] Test {t_id} BAŞARISIZ (Status: {target_channel_status})\n")

            self.test_results[t_id] = {
                "id": t_id, "name": t_name, "type": test["type"],
                "expected": test["expected"], "status": status_str,
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }

        except Exception as e:
            self.log(f"    [SİSTEM HATASI] İstisna oluştu: {e}\n")

    def run_single_test(self, test_id):
        self.com_port = self.port_entry.get()
        test = next((t for t in self.test_scenarios if t["id"] == test_id), None)
        if not test:
            self.set_buttons_state("normal")
            return

        client = ModbusSerialClient(
            port=self.com_port, baudrate=115200, parity='N', stopbits=1, bytesize=8, timeout=1
        )

        if not client.connect():
            self.log(f"[HATA] {self.com_port} portuna bağlanılamadı!")
            self.set_buttons_state("normal")
            return

        self.execute_modbus_test(client, test)
        client.close()

        self.set_buttons_state("normal")
        self.enable_report_buttons()

    def run_all_tests(self):
        self.com_port = self.port_entry.get()
        self.log_box.delete("1.0", "end")
        self.test_results.clear()

        client = ModbusSerialClient(
            port=self.com_port, baudrate=115200, parity='N', stopbits=1, bytesize=8, timeout=1
        )

        if not client.connect():
            self.log(f"[HATA] {self.com_port} portuna bağlanılamadı!")
            self.set_buttons_state("normal")
            return

        for test in self.test_scenarios:
            self.execute_modbus_test(client, test)

        client.close()
        self.set_buttons_state("normal")
        self.enable_report_buttons()

    def generate_excel_report(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "HIL Test Raporu"

        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        font_pass = Font(name="Calibri", size=11, bold=True, color="006100")
        fill_pass = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        font_fail = Font(name="Calibri", size=11, bold=True, color="9C0006")
        fill_fail = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )

        ws.merge_cells("A1:F1")
        ws["A1"] = "Taytech TTSimBox - 8-Kanal HIL Doğrulama Raporu"
        ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
        ws["A1"].alignment = align_left
        ws["A2"] = f"Rapor Tarihi: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(name="Calibri", size=10, italic=True)

        headers = ["Test ID", "Senaryo Adı", "Test Tipi", "Beklenen Değer", "Test Durumu", "Zaman Damgası"]
        ws.append([])
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border

        row_idx = 5
        for t_id in sorted(self.test_results.keys()):
            r = self.test_results[t_id]
            row_data = [r["id"], r["name"], r["type"], r["expected"], r["status"], r["timestamp"]]
            ws.append(row_data)

            for col_num in range(1, 7):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.border = thin_border
                cell.alignment = align_center if col_num in [1, 3, 4, 5, 6] else align_left
                if col_num == 5:
                    if r["status"] == "PASSED":
                        cell.font = font_pass
                        cell.fill = fill_pass
                    else:
                        cell.font = font_fail
                        cell.fill = fill_fail
            row_idx += 1

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"HIL_8Channel_Report_{timestamp_str}.xlsx"
        wb.save(excel_filename)
        self.log(f"\n[EXCEL] Rapor oluşturuldu: {excel_filename}")
        os.system(f"start {excel_filename}")

    def generate_pdf_report(self):
        pdf_filename = f"HIL_8Channel_Report_{int(time.time())}.pdf"
        doc = SimpleDocTemplate(pdf_filename, pagesize=letter)
        styles = getSampleStyleSheet()

        story = []
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, leading=22, textColor=colors.HexColor("#1A365D"))
        story.append(Paragraph("TTSimBox 8-Channel HIL Test Report", title_style))
        story.append(Spacer(1, 15))

        data = [["Test ID", "Senaryo Adı", "Tip", "Beklenen Değer", "Durum"]]
        for t_id in sorted(self.test_results.keys()):
            r = self.test_results[t_id]
            data.append([str(r["id"]), r["name"], r["type"], r["expected"], r["status"]])

        t = Table(data, colWidths=[50, 240, 110, 80, 70])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#2B6CB0")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey)
        ]))

        story.append(t)
        doc.build(story)
        self.log(f"\n[PDF] Rapor oluşturuldu: {pdf_filename}")
        os.system(f"start {pdf_filename}")


if __name__ == "__main__":
    app = HILTestApp()
    app.mainloop()